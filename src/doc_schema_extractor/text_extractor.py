"""Document text and layout extraction - pdfplumber (MIT) + openpyxl (MIT)."""

from __future__ import annotations

import re
import warnings
from pathlib import Path
from dataclasses import dataclass, field

import openpyxl
import pdfplumber

from .logging_utils import get_logger

logger = get_logger("text_extractor")

_RE_LOWER_UPPER = re.compile(r'([a-z])([A-Z])')
_RE_DIGIT_ALPHA = re.compile(r'(\d)([A-Za-z])')
_RE_ALPHA_DIGIT = re.compile(r'([A-Za-z])(\d)')
_RE_MULTI_SPACE = re.compile(r'  +')
# Strip [Sheet: ...] markers before fingerprint normalisation so sheet names
# never become discriminating keywords.
_RE_SHEET_MARKER = re.compile(r'\[Sheet:[^\]]*\]\s*')


def normalise_text(text: str) -> str:
    """Insert spaces at camelCase and digit/letter boundaries.

    Also strips [Sheet: ...] markers so XLSX sheet names do not pollute
    fingerprint keyword matching.
    """
    text = _RE_SHEET_MARKER.sub(' ', text)
    text = _RE_LOWER_UPPER.sub(r'\1 \2', text)
    text = _RE_DIGIT_ALPHA.sub(r'\1 \2', text)
    text = _RE_ALPHA_DIGIT.sub(r'\1 \2', text)
    text = _RE_MULTI_SPACE.sub(' ', text)
    return text


@dataclass
class PageContent:
    page_num: int
    text: str
    tables: list[list[list[str | None]]] = field(default_factory=list)
    sheet_name: str = ""


@dataclass
class DocumentContent:
    path: str
    file_type: str
    full_text: str
    normalised_text: str
    pages: list[PageContent] = field(default_factory=list)
    sheet_names: list[str] = field(default_factory=list)


class TextExtractor:
    def extract(self, file_path: str | Path) -> DocumentContent:
        path = Path(file_path)
        logger.debug("Starting extraction for file=%s suffix=%s", path, path.suffix.lower())
        if not path.exists():
            logger.error("File not found: %s", file_path)
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = path.suffix.lower()
        if suffix == ".pdf":
            return self._extract_pdf(path)
        if suffix in (".xlsx", ".xls"):
            return self._extract_xlsx(path)

        logger.error("Unsupported file type: %s", suffix)
        raise ValueError(f"Unsupported file type: {suffix}")

    def _extract_pdf(self, path: Path) -> DocumentContent:
        pages: list[PageContent] = []
        all_text_parts: list[str] = []
        logger.debug("Extracting PDF path=%s", path)

        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text(x_tolerance=3, y_tolerance=3) or ""
                tables = page.extract_tables() or []
                normalized_tables = [
                    [[str(cell) if cell is not None else "" for cell in row] for row in table]
                    for table in tables
                ]
                logger.debug(
                    "PDF page=%s chars=%s tables=%s",
                    i + 1, len(text), len(normalized_tables),
                )
                pages.append(PageContent(page_num=i + 1, text=text, tables=normalized_tables))
                all_text_parts.append(text)

        full_text = "\n".join(all_text_parts)
        logger.info("Finished PDF extraction path=%s pages=%s", path, len(pages))
        return DocumentContent(
            path=str(path),
            file_type="pdf",
            full_text=full_text,
            normalised_text=normalise_text(full_text),
            pages=pages,
        )

    def _extract_xlsx(self, path: Path) -> DocumentContent:
        logger.debug("Extracting XLSX path=%s", path)

        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message="Cannot parse header or footer",
                category=UserWarning,
                module=r"openpyxl",
            )
            warnings.filterwarnings(
                "ignore",
                message="Data Validation extension is not supported",
                category=UserWarning,
                module=r"openpyxl",
            )
            wb = openpyxl.load_workbook(path, data_only=True)

        pages: list[PageContent] = []
        all_text_parts: list[str] = []
        sheet_names: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            text_lines: list[str] = []

            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(cells)
                    text_lines.append("  ".join(cells))

            text = "\n".join(text_lines)
            logger.debug("XLSX sheet=%s rows=%s chars=%s", sheet_name, len(rows), len(text))
            pages.append(PageContent(
                page_num=len(pages) + 1, text=text, tables=[rows], sheet_name=sheet_name
            ))
            sheet_names.append(sheet_name)
            # [Sheet: ...] markers kept in full_text for rule-engine anchoring,
            # but stripped from normalised_text by normalise_text() above.
            all_text_parts.append(f"[Sheet: {sheet_name}]\n{text}")

        full_text = "\n".join(all_text_parts)
        logger.info(
            "Finished XLSX extraction path=%s sheets=%s names=%s",
            path, len(pages), sheet_names,
        )
        return DocumentContent(
            path=str(path),
            file_type="xlsx",
            full_text=full_text,
            normalised_text=normalise_text(full_text),
            pages=pages,
            sheet_names=sheet_names,
        )
